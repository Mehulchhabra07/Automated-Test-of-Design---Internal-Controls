"""
------------------------------------------------------------------
AI-Powered Test of Design (TOD) Control Analysis Tool
requires : pandas  openpyxl  openai  httpx
------------------------------------------------------------------
An intelligent auditing framework that uses AI to analyze internal controls 
for completeness and effectiveness. Features automated risk assessment, 
control evaluation, and comprehensive reporting with Excel output.

This tool automates the traditionally manual process of evaluating internal
controls, reducing analysis time from hours to minutes while maintaining
high accuracy and consistency.

Author: Mehul Chhabra
GitHub: https://github.com/Mehulchhabra07/Automated-Test-of-Design---Internal-Controls
Project: AI-Driven Control Testing Framework
------------------------------------------------------------------
"""

from pathlib import Path
from datetime import datetime
import os, sys, json, time, re, logging
import pandas as pd
import httpx
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openai import OpenAI, OpenAIError
from typing import Tuple, Optional, Dict, Any

# ═══════════════════════════════════════════════════════════════════════════════
#                               CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

class Config:
    """Configuration settings for TOD Control Testing"""
    
    # File locations - Update these paths for your environment
    INPUT_FILE = Path("sample_controls.xlsx")  # Place your input file in the same directory
    OUTPUT_FILE = INPUT_FILE.with_name(f"{INPUT_FILE.stem}_TestResult.xlsx")
    
    # Input columns (exact headers in row‑1)
    CONTROL_COL = "Control"
    DESC_COL = "Control Description"
    
    # Completeness elements the LLM must assess
    ELEMENTS = ["When", "Why", "Who", "What", "Where", "How"]
    
    # OpenAI API settings - Set your API key here or via environment variable
    API_KEY = os.getenv("OPENAI_API_KEY", "YOUR_OPENAI_API_KEY_HERE")
    BASE_URL = "https://api.openai.com/v1"  # Standard OpenAI API endpoint
    MODEL = "gpt-4"  # Using GPT-4 for better analysis quality
    
    # Supported OpenAI models
    SUPPORTED_MODELS = [
        "gpt-4", "gpt-4-turbo", "gpt-3.5-turbo", 
        "gpt-4o", "gpt-4o-mini"
    ]
    
    # Retry settings for robust error handling
    MAX_RETRIES = 5
    RETRY_DELAY = 1.0
    MAX_RETRY_DELAY = 60.0
    
    # Request timeout settings
    REQUEST_TIMEOUT = 120.0
    
    # Required columns for validation
    REQUIRED_COLS = [
        "Risk", "Risk Description", "Control", "Control Description",
        "Automation", "Detective/ Preventive", "Operation Frequency"
    ]

# ═══════════════════════════════════════════════════════════════════════════════
#                               LOGGING SETUP
# ═══════════════════════════════════════════════════════════════════════════════

def setup_logging():
    """Setup logging configuration"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('tod_analysis.log'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logging()

# ═══════════════════════════════════════════════════════════════════════════════
#                               API CLIENT SETUP
# ═══════════════════════════════════════════════════════════════════════════════

def initialize_client() -> OpenAI:
    """Initialize and test OpenAI client with robust error handling"""
    if Config.API_KEY == "YOUR_OPENAI_API_KEY_HERE":
        logger.error("⚠  Set OPENAI_API_KEY environment variable or update Config.API_KEY")
        sys.exit(1)

    # Validate model name
    if Config.MODEL not in Config.SUPPORTED_MODELS:
        logger.warning(f"⚠ Model '{Config.MODEL}' not in supported models list: {Config.SUPPORTED_MODELS}")
        logger.info(f"Continuing with '{Config.MODEL}' - it may still work")

    logger.info(f"Initializing OpenAI client with model: {Config.MODEL}")
    logger.info(f"Base URL: {Config.BASE_URL}")

    client = OpenAI(
        api_key=Config.API_KEY,
        base_url=Config.BASE_URL,
        http_client=httpx.Client(
            verify=True,  # Enable SSL verification for standard OpenAI API
            timeout=Config.REQUEST_TIMEOUT
        )
    )

    # Test connection with retry logic
    for attempt in range(Config.MAX_RETRIES):
        try:
            logger.info(f"Testing connection (attempt {attempt + 1}/{Config.MAX_RETRIES})...")
            response = client.chat.completions.create(
                model=Config.MODEL,
                messages=[
                    {"role": "system", "content": "Be concise and precise."},
                    {"role": "user", "content": "ping"}
                ],
                max_tokens=10
            )
            
            if response and response.choices and response.choices[0].message.content:
                logger.info("✔ OpenAI API connection test successful")
                logger.info(f"✔ Model '{Config.MODEL}' is working correctly")
                return client
            else:
                raise Exception("Invalid response structure from API")
                
        except Exception as e:
            error_msg = str(e).lower()
            logger.warning(f"Connection test attempt {attempt + 1} failed: {e}")
            
            # Handle specific error types
            if "429" in error_msg or "too many requests" in error_msg:
                wait_time = min(Config.RETRY_DELAY * (2 ** attempt), Config.MAX_RETRY_DELAY)
                logger.info(f"Rate limit hit, waiting {wait_time}s before retry...")
                time.sleep(wait_time)
            elif "401" in error_msg or "unauthorized" in error_msg:
                logger.error("🚫 Authentication failed - check your API key")
                sys.exit(1)
            elif "404" in error_msg or "not found" in error_msg:
                logger.error("🚫 API endpoint not found - check base URL and model name")
                sys.exit(1)
            elif attempt < Config.MAX_RETRIES - 1:
                wait_time = Config.RETRY_DELAY * (2 ** attempt)
                logger.info(f"Retrying in {wait_time}s...")
                time.sleep(wait_time)
            else:
                logger.error(f"🚫 All connection attempts failed. Last error: {e}")
                logger.error("Please check:")
                logger.error("1. API key is valid and not expired")
                logger.error("2. Model name is supported")
                logger.error("3. Network connectivity")
                logger.error("4. OpenAI API service status")
                sys.exit(1)

    logger.error("🚫 OpenAI API connection could not be established")
    sys.exit(1)

# ═══════════════════════════════════════════════════════════════════════════════
#                               UTILITY FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

BRACE_RE = re.compile(r"\{.*\}", re.S)

def extract_json_from_response(raw_response: str) -> Optional[Dict[str, Any]]:
    """Extract JSON from LLM response with robust error handling"""
    try:
        # Clean up response
        if raw_response.lower().startswith("json"):
            raw_response = raw_response[4:].strip()
        
        # Try to find JSON in response
        match = BRACE_RE.search(raw_response)
        json_str = match.group(0) if match else raw_response
        
        return json.loads(json_str)
    except json.JSONDecodeError as e:
        logger.warning(f"Failed to parse JSON response: {e}")
        return None
    except Exception as e:
        logger.warning(f"Unexpected error parsing response: {e}")
        return None

def make_llm_call_with_retry(client: OpenAI, prompt: str, system_content: str = "Respond only with the JSON object.") -> Optional[str]:
    """Make LLM call with retry logic and error handling"""
    for attempt in range(Config.MAX_RETRIES):
        try:
            response = client.chat.completions.create(
                model=Config.MODEL,
                messages=[
                    {"role": "system", "content": system_content},
                    {"role": "user", "content": prompt}
                ],
            )
            return response.choices[0].message.content.strip()
        except Exception as e:
            error_msg = str(e).lower()
            logger.warning(f"LLM call attempt {attempt + 1} failed: {e}")
            
            # Handle specific error types
            if "429" in error_msg or "too many requests" in error_msg:
                wait_time = min(Config.RETRY_DELAY * (2 ** attempt), Config.MAX_RETRY_DELAY)
                logger.info(f"Rate limit hit, waiting {wait_time}s before retry...")
                time.sleep(wait_time)
            elif "401" in error_msg or "unauthorized" in error_msg:
                logger.error("🚫 Authentication failed during LLM call")
                return None
            elif "404" in error_msg or "not found" in error_msg:
                logger.error("🚫 Model or endpoint not found during LLM call")
                return None
            elif attempt < Config.MAX_RETRIES - 1:
                wait_time = Config.RETRY_DELAY * (2 ** attempt)  # Exponential backoff
                logger.info(f"Retrying LLM call in {wait_time}s...")
                time.sleep(wait_time)
            else:
                logger.error(f"All {Config.MAX_RETRIES} LLM call attempts failed")
    return None

# ═══════════════════════════════════════════════════════════════════════════════
#                               ANALYSIS FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def ask_llm(client: OpenAI, text: str) -> Tuple[str, str, str]:
    """
    Core AI analysis function that evaluates control descriptions for completeness.
    
    This function implements a sophisticated prompt engineering approach to analyze
    internal controls against the 6W framework (Who, What, When, Where, Why, How).
    
    Args:
        client: OpenAI client instance
        text: Control description to analyze
        
    Returns:
        Tuple containing (present_elements, missing_elements, suggestions)
    """
    prompt = f"""
You are the world's best professional auditor with decades of experience testing control descriptions for completeness.

Given the following control description, evaluate it for the presence of six key elements: {", ".join(Config.ELEMENTS)}.

For each element:
- If present, list it with a short clause (<20 words) referencing how it's reflected in the description.
- If missing, explain briefly why it's considered missing (e.g., "no timeline or frequency mentioned").

Then suggest improvements for each missing element based on the description.

Return valid JSON in this format:
{{
  "present": {{
    "Who": "...",
    "What": "...",
    ...
  }},
  "missing": {{
    "When": "No timeline stated",
    "Where": "No tool or system mentioned",
    ...
  }},
  "suggestions": {{
    "When": "Suggest adding a specific timeline or frequency for review",
    ...
  }}
}}

Control Description:
\"\"\"{text}\"\"\"

JSON:
"""

    raw_response = make_llm_call_with_retry(client, prompt)
    if not raw_response:
        return "", "LLM error", ""
    
    data = extract_json_from_response(raw_response)
    if not data:
        return "", "LLM error", ""

    present = data.get("present", {})
    missing = data.get("missing", {})
    suggestions = data.get("suggestions", {})

    present_str = "\n".join([f"• {k}: {v}" for k, v in present.items()])
    missing_str = "\n".join([f"• {k}: {v}" for k, v in missing.items()])
    suggestions_str = "\n".join([f"• {k}: {v}" for k, v in suggestions.items()])

    return present_str, missing_str, suggestions_str

def ask_control_objective(client: OpenAI, risk_desc: str, control_desc: str) -> Tuple[str, str]:
    """Assess if control is designed to mitigate the risk"""
    prompt = f"""
Given the following risk description and control description, answer:
1. Is the control, as designed, able to mitigate the risk? (Yes/No)
2. Briefly explain your reasoning (1-2 sentences).

Risk Description:
{risk_desc}

Control Description:
{control_desc}

Respond in JSON:
{{
  "answer": "Yes or No",
  "explanation": "..."
}}
"""
    raw_response = make_llm_call_with_retry(client, prompt)
    if not raw_response:
        return "LLM error", "LLM error"
    
    data = extract_json_from_response(raw_response)
    if not data:
        return "LLM error", "LLM error"
    
    return data.get("answer", "LLM error"), data.get("explanation", "LLM error")

def ask_execution_appropriateness(client: OpenAI, automation: str, risk_desc: str, control_desc: str) -> Tuple[str, str]:
    """Assess execution appropriateness"""
    prompt = f"""
Given the automation type (Automated/Semi-Auto/Manual), risk description, and control description, answer:
1. Is the control execution appropriate based on the control description and risk description? (Yes/No)
2. Briefly explain your reasoning (1-2 sentences).

Automation: {automation}
Risk Description: {risk_desc}
Control Description: {control_desc}

Respond in JSON:
{{
  "answer": "Yes or No",
  "explanation": "..."
}}
"""
    raw_response = make_llm_call_with_retry(client, prompt)
    if not raw_response:
        return "LLM error", "LLM error"
    
    data = extract_json_from_response(raw_response)
    if not data:
        return "LLM error", "LLM error"
    
    return data.get("answer", "LLM error"), data.get("explanation", "LLM error")

def ask_type_adequacy(client: OpenAI, control_type: str, risk_desc: str, control_desc: str) -> Tuple[str, str]:
    """Assess control type adequacy"""
    prompt = f"""
Given the control type (Detective/Preventive), risk description, and control description, answer:
1. Is the control type appropriate based on control description and adequate for the risk it addresses? (Yes/No)
2. Briefly explain your reasoning (1-2 sentences).

Type: {control_type}
Risk Description: {risk_desc}
Control Description: {control_desc}

Respond in JSON:
{{
  "answer": "Yes or No",
  "explanation": "..."
}}
"""
    raw_response = make_llm_call_with_retry(client, prompt)
    if not raw_response:
        return "LLM error", "LLM error"
    
    data = extract_json_from_response(raw_response)
    if not data:
        return "LLM error", "LLM error"
    
    return data.get("answer", "LLM error"), data.get("explanation", "LLM error")

def ask_frequency_appropriateness(client: OpenAI, frequency: str, risk_desc: str, control_desc: str) -> Tuple[str, str]:
    """Assess frequency appropriateness"""
    prompt = f"""
Given the operation frequency, risk description, and control description, answer:
1. Is the control frequency appropriate based on the control description and adequated for the associated risk? (Yes/No)
2. Briefly explain your reasoning (1-2 sentences).

Frequency: {frequency}
Risk Description: {risk_desc}
Control Description: {control_desc}

Respond in JSON:
{{
  "answer": "Yes or No",
  "explanation": "..."
}}
"""
    raw_response = make_llm_call_with_retry(client, prompt)
    if not raw_response:
        return "LLM error", "LLM error"
    
    data = extract_json_from_response(raw_response)
    if not data:
        return "LLM error", "LLM error"
    
    return data.get("answer", "LLM error"), data.get("explanation", "LLM error")

def ask_system_dependency(client: OpenAI, control_desc: str) -> Tuple[str, str]:
    """Extract system dependencies"""
    prompt = f"""
Given the control description, extract the names of any systems or data sources mentioned. List only the system or data source names (comma-separated if more than one). If none are mentioned, return "None found".

Control Description: {control_desc}

Respond in JSON:
{{
  "systems": "..."
}}
"""
    raw_response = make_llm_call_with_retry(client, prompt)
    if not raw_response:
        return "LLM error", ""
    
    data = extract_json_from_response(raw_response)
    if not data:
        return "LLM error", ""
    
    return data.get("systems", "LLM error"), ""

def ask_adaptability(client: OpenAI, control_desc: str) -> Tuple[str, str]:
    """Assess control SOD"""
    prompt = f"""
Given the following control description, answer:
1. Does the control ensure that no single individual has end-to-end responsibility for critical transactions so like proper Segregation of duties? (Yes/No)
2. Briefly explain your reasoning (1-2 sentences).

Control Description:
{control_desc}

Respond in JSON:
{{
  "answer": "Yes or No",
  "explanation": "..."
}}
"""
    raw_response = make_llm_call_with_retry(client, prompt)
    if not raw_response:
        return "LLM error", "LLM error"
    
    data = extract_json_from_response(raw_response)
    if not data:
        return "LLM error", "LLM error"
    
    return data.get("answer", "LLM error"), data.get("explanation", "LLM error")

def ask_overall_rating(client: OpenAI, row: Dict[str, str], present: str, missing: str, adaptability: str) -> Tuple[str, str]:
    """Provide overall control rating"""
    prompt = f"""
Given the following analysis of a control, provide an overall rating as one of the following: Effective, Partially effective, In-effective. Consider all the information below:
- Control objective: {row['Control objective: Is the control designed able to mitigate the risk ?']}
- Execution appropriateness: {row['Is the control execution appropriate for the risk being addressed?']}
- Type adequacy: {row['Is the control type adequate for the risk it addresses']}
- Frequency appropriateness: {row['Is the control frequency appropriate for the associated risk?']}
- System/data dependencies: {row['System/data dependencies: Are the systems/data sources used reliable and secure?']}
- Present: {present}
- Missing: {missing}

Return a JSON object:
{{
  "rating": "Effective, Partially effective, or In-effective",
  "explanation": "..."
}}
"""
    raw_response = make_llm_call_with_retry(client, prompt)
    if not raw_response:
        return "LLM error", "LLM error"
    
    data = extract_json_from_response(raw_response)
    if not data:
        return "LLM error", "LLM error"
    
    return data.get("rating", "LLM error"), data.get("explanation", "LLM error")

def ask_expected_evidence(client: OpenAI, control_desc: str) -> str:
    """Generate expected evidence list"""
    prompt = f"""
You are the world's best professional auditor with decades of experience testing control descriptions for completeness.

Given the following control description, list the types of evidence an auditor or tester would expect to see to verify the control's operation. List each expected evidence as a separate numbered point (1., 2., 3., etc.).

Control Description:
{control_desc}

Respond with a numbered list of expected evidence types only.
"""
    raw_response = make_llm_call_with_retry(client, prompt, "Respond only with the numbered list.")
    return raw_response.strip() if raw_response else "LLM error"

# ═══════════════════════════════════════════════════════════════════════════════
#                               DATA PROCESSING
# ═══════════════════════════════════════════════════════════════════════════════

def load_and_validate_data() -> pd.DataFrame:
    """Load Excel file and validate required columns"""
    try:
        logger.info(f"Loading data from {Config.INPUT_FILE}")
        df = pd.read_excel(Config.INPUT_FILE, engine="openpyxl")
        
        # Validate required columns
        missing_cols = [col for col in Config.REQUIRED_COLS if col not in df.columns]
        if missing_cols:
            raise ValueError(f"Missing required columns: {missing_cols}")
        
        logger.info(f"✔ Loaded {len(df)} controls with all required columns")
        return df
        
    except Exception as e:
        logger.error(f"⚠ Could not read {Config.INPUT_FILE} → {e}")
        sys.exit(1)

def process_controls(client: OpenAI, df: pd.DataFrame) -> pd.DataFrame:
    """Process all controls and generate analysis results"""
    logger.info("Starting control analysis...")
    
    # Initialize result lists
    present_list, missing_list, suggestion_list = [], [], []
    objective_ans, objective_exp = [], []
    exec_ans, exec_exp = [], []
    type_ans, type_exp = [], []
    freq_ans, freq_exp = [], []
    sysdep_ans, sysdep_exp = [], []
    present_missing_list = []
    adaptability_ans, adaptability_exp = [], []
    overall_rating_ans, overall_rating_exp = [], []
    potential_evidence_list = []

    total_controls = len(df)
    
    for idx, row in df.iterrows():
        control_name = row.get("Control", f"Control_{idx+1}")
        logger.info(f"Processing [{idx+1}/{total_controls}]: {control_name}")
        
        try:
            # 1. Completeness analysis
            pres, miss, sugg = ask_llm(client, row["Control Description"])
            present_list.append(pres)
            missing_list.append(miss)
            suggestion_list.append(sugg)
            present_missing = f"Present:\n{pres}\n\nMissing:\n{miss}"
            present_missing_list.append(present_missing)
            
            # 2. Control objective
            ans, exp = ask_control_objective(client, row["Risk Description"], row["Control Description"])
            objective_ans.append(ans)
            objective_exp.append(exp)
            
            # 3. Execution appropriateness
            ans2, exp2 = ask_execution_appropriateness(client, row["Automation"], row["Risk Description"], row["Control Description"])
            exec_ans.append(ans2)
            exec_exp.append(exp2)
            
            # 4. Type adequacy
            ans3, exp3 = ask_type_adequacy(client, row["Detective/ Preventive"], row["Risk Description"], row["Control Description"])
            type_ans.append(ans3)
            type_exp.append(exp3)
            
            # 5. Frequency appropriateness
            ans4, exp4 = ask_frequency_appropriateness(client, row["Operation Frequency"], row["Risk Description"], row["Control Description"])
            freq_ans.append(ans4)
            freq_exp.append(exp4)
            
            # 6. System/data dependencies
            ans5, exp5 = ask_system_dependency(client, row["Control Description"])
            sysdep_ans.append(ans5)
            sysdep_exp.append(exp5)
            
            # 7. Adaptability
            ans6, exp6 = ask_adaptability(client, row["Control Description"])
            adaptability_ans.append(ans6)
            adaptability_exp.append(exp6)
            
            # 8. Overall Rating
            overall_row = {
                'Control objective: Is the control designed able to mitigate the risk ?': ans,
                'Is the control execution appropriate for the risk being addressed?': ans2,
                'Is the control type adequate for the risk it addresses': ans3,
                'Is the control frequency appropriate for the associated risk?': ans4,
                'System/data dependencies: Are the systems/data sources used reliable and secure?': ans5
            }
            ans7, exp7 = ask_overall_rating(client, overall_row, pres, miss, ans6)
            overall_rating_ans.append(ans7)
            overall_rating_exp.append(exp7)
            
            # 9. Expected evidence
            expected_evidence = ask_expected_evidence(client, row["Control Description"])
            potential_evidence_list.append(expected_evidence)
            
            logger.info(f"  ✓ Completed analysis for {control_name}")
            
        except Exception as e:
            logger.error(f"  ✗ Error processing {control_name}: {e}")
            # Add error placeholders to maintain data consistency
            for lst in [present_list, missing_list, suggestion_list, present_missing_list,
                       objective_ans, objective_exp, exec_ans, exec_exp, type_ans, type_exp,
                       freq_ans, freq_exp, sysdep_ans, sysdep_exp, adaptability_ans, adaptability_exp,
                       overall_rating_ans, overall_rating_exp, potential_evidence_list]:
                if len(lst) <= idx:
                    lst.append("Processing error")

    # Add all results to DataFrame (maintaining exact same column structure)
    df["Present & Missing"] = present_missing_list
    
    # Remove columns if they exist (same as original)
    for col in ["Present", "Missing"]:
        if col in df.columns:
            del df[col]
    
    df["Suggestions"] = suggestion_list
    
    # Remove Source column if it exists
    if "Source" in df.columns:
        del df["Source"]
    
    # Add all analysis columns (same order as original)
    df["Control objective: Is the control designed able to mitigate the risk ?"] = objective_ans
    df["Control objective: Explanation"] = objective_exp
    df["Is the control execution appropriate for the risk being addressed?"] = exec_ans
    df["Execution appropriateness: Explanation"] = exec_exp
    df["Is the control type adequate for the risk it addresses"] = type_ans
    df["Type adequacy: Explanation"] = type_exp
    df["Is the control frequency appropriate for the associated risk?"] = freq_ans
    df["Frequency appropriateness: Explanation"] = freq_exp
    df["System/data dependencies: Are the systems/data sources used reliable and secure?"] = sysdep_ans
    df["Adaptability - Is the control adaptable to new risks or process changes?"] = adaptability_ans
    df["Adaptability: Explanation"] = adaptability_exp
    df["Overall Rating"] = overall_rating_ans
    df["Overall Rating: Explanation"] = overall_rating_exp
    df["Potential Evidences Expected Based on Control Description"] = potential_evidence_list

    # Remove System/data dependencies: Explanation column if it exists
    if "System/data dependencies: Explanation" in df.columns:
        del df["System/data dependencies: Explanation"]

    # Update Present & Missing column header (same as original)
    present_missing_col = "Has the control been formally documented? (When, Why, Who, What, Where and How)"
    if "Present & Missing" in df.columns:
        df.rename(columns={"Present & Missing": present_missing_col}, inplace=True)

    logger.info("✓ Control analysis completed")
    return df

# ═══════════════════════════════════════════════════════════════════════════════
#                               EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════

def save_results_to_excel(df: pd.DataFrame):
    """Save results to Excel with formatting (identical to original)"""
    logger.info(f"Saving results to {Config.OUTPUT_FILE}")
    
    try:
        with pd.ExcelWriter(Config.OUTPUT_FILE, engine="openpyxl") as xl:
            df.to_excel(xl, sheet_name="TOD Results", index=False, startrow=2)

            wb = xl.book
            ws = wb["TOD Results"]

            # Insert two rows at the top
            ws.insert_rows(1, amount=2)
            
            # Merge and label input columns
            input_cols = [
                "Risk", "Risk Description", "Control", "Control Description",
                "Automation", "Detective/ Preventive", "Operation Frequency"
            ]
            output_cols = [col for col in df.columns if col not in input_cols and col != "Source"]
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(input_cols))
            ws.cell(row=1, column=1).value = "INPUT COLUMNS"
            ws.merge_cells(start_row=1, start_column=len(input_cols)+1, end_row=1, end_column=len(input_cols)+len(output_cols))
            ws.cell(row=1, column=len(input_cols)+1).value = "OUTPUT COLUMNS"

            # Remove Source column if present
            if "Source" in [cell.value for cell in ws[3] if cell.value]:
                try:
                    idx = [cell.value for cell in ws[3]].index("Source") + 1
                    ws.delete_cols(idx)
                except (ValueError, IndexError):
                    pass  # Source column not found, ignore

            # Set alignment and wrap text for all cells
            align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = align

            # Apply formatting to the Present & Missing column (identical to original)
            from openpyxl.styles import Font, PatternFill
            present_missing_col = "Has the control been formally documented? (When, Why, Who, What, Where and How)"
            present_missing_idx = None
            for idx, cell in enumerate(ws[3], 1):
                if cell.value == present_missing_col:
                    present_missing_idx = idx
                    break
            if present_missing_idx:
                for row in ws.iter_rows(min_row=4, min_col=present_missing_idx, max_col=present_missing_idx, max_row=ws.max_row):
                    cell = row[0]
                    value = cell.value or ""
                    # Try to mimic rich formatting (same logic as original)
                    lines = value.split("\n")
                    new_lines = []
                    for line in lines:
                        if line.strip().startswith("Present:"):
                            new_lines.append("**Present:**" + line[len("Present:"):])
                        elif line.strip().startswith("Missing:"):
                            new_lines.append("**Missing:**" + line[len("Missing:"):])
                        elif line.strip().startswith("•"):
                            # Bold element name (before ':')
                            if ":" in line:
                                elem, rest = line.split(":", 1)
                                elem = elem.strip()
                                # If in missing section, color red
                                if any("Missing" in l for l in new_lines[-2:]):
                                    # Use red font for whole line
                                    new_lines.append(f"[RED]{elem}:[/RED]{rest}")
                                else:
                                    new_lines.append(f"**{elem}:**{rest}")
                            else:
                                new_lines.append(line)
                        else:
                            new_lines.append(line)
                    # Compose new value with formatting hints
                    cell.value = "\n".join(new_lines)
                    # Make whole cell bold for Present & Missing
                    cell.font = Font(bold=True)
                    # If any [RED] marker, color cell red (simulate missing points)
                    if any("[RED]" in l for l in new_lines):
                        cell.font = Font(bold=True, color="FF0000")
                    else:
                        cell.font = Font(bold=True)

            # Auto-width (same as original)
            for col_idx, _ in enumerate(ws[3], 1):
                max_len = max(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

            wb.save(Config.OUTPUT_FILE)

        logger.info(f"✓ Results saved successfully to {Config.OUTPUT_FILE}")
        
    except Exception as e:
        logger.error(f"Error saving Excel file: {e}")
        raise

# ═══════════════════════════════════════════════════════════════════════════════
#                               MAIN EXECUTION
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    """Main execution function"""
    start_time = time.time()
    
    try:
        logger.info("="*80)
        logger.info("AI-Powered TOD Control Testing Framework")
        logger.info("="*80)
        
        # Initialize client
        client = initialize_client()
        
        # Load and validate data
        df = load_and_validate_data()
        
        # Process controls
        df_results = process_controls(client, df)
        
        # Save results
        save_results_to_excel(df_results)
        
        # Summary
        elapsed_time = time.time() - start_time
        logger.info("="*80)
        logger.info(f"✓ Analysis completed successfully!")
        logger.info(f"✓ Processed {len(df)} controls in {elapsed_time:.1f} seconds")
        logger.info(f"✓ Results saved to: {Config.OUTPUT_FILE}")
        logger.info("="*80)
        
    except KeyboardInterrupt:
        logger.info("\n⚠ Analysis interrupted by user")
        sys.exit(1)
    except Exception as e:
        logger.error(f"✗ Analysis failed: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
